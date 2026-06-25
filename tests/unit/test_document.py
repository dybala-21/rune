"""Tests for the document_create capability (xlsx/pptx/docx/pdf/csv/html)."""

from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import pytest

from rune.capabilities.document import (
    DocBlock,
    DocSheet,
    DocumentCreateParams,
    DocumentReadParams,
    document_create,
    document_read,
)


@pytest.fixture
def tmp_path():
    # The Guardian blocks writes under /var (pytest's default tmp), so use a
    # HOME-based directory the safety layer allows.
    d = Path.home() / ".rune_doctest"
    d.mkdir(parents=True, exist_ok=True)
    try:
        yield d
    finally:
        shutil.rmtree(d, ignore_errors=True)

_BLOCKS = [
    DocBlock(type="heading", text="Overview", level=1),
    DocBlock(type="paragraph", text="Local-first AI."),
    DocBlock(type="bullets", items=["Private", "Verified", "Fast"]),
    DocBlock(type="table", rows=[["Product", "Qty"], ["Laptop", 15]]),
    DocBlock(type="page_break"),
    DocBlock(type="paragraph", text="Second page."),
]
_SHEETS = [DocSheet(name="Sales", rows=[["Product", "Qty"], ["Laptop", 15]])]


@pytest.mark.asyncio
@pytest.mark.parametrize("fmt", ["xlsx", "csv", "docx", "pptx", "pdf", "html"])
async def test_each_format_renders(tmp_path, fmt):
    out = tmp_path / f"doc.{fmt}"
    r = await document_create(DocumentCreateParams(
        path=str(out), format=fmt, title="Report",
        blocks=_BLOCKS, sheets=_SHEETS,
    ))
    assert r.success, r.error
    assert out.is_file() and out.stat().st_size > 0


@pytest.mark.asyncio
async def test_ooxml_outputs_are_valid_zips(tmp_path):
    for fmt in ("xlsx", "docx", "pptx"):
        out = tmp_path / f"doc.{fmt}"
        r = await document_create(DocumentCreateParams(
            path=str(out), format=fmt, title="R", blocks=_BLOCKS, sheets=_SHEETS,
        ))
        assert r.success
        assert zipfile.is_zipfile(out), f"{fmt} is not a valid OOXML zip"


@pytest.mark.asyncio
async def test_xlsx_preserves_numeric_cells(tmp_path):
    from openpyxl import load_workbook

    out = tmp_path / "n.xlsx"
    await document_create(DocumentCreateParams(
        path=str(out), format="xlsx",
        sheets=[DocSheet(name="S", rows=[["k", "v"], ["a", 12345]])],
    ))
    ws = load_workbook(out)["S"]
    assert ws["B2"].value == 12345
    assert isinstance(ws["B2"].value, int)


@pytest.mark.asyncio
@pytest.mark.parametrize("fmt", ["xlsx", "docx", "pptx", "pdf", "csv", "html"])
async def test_create_read_roundtrip(tmp_path, fmt):
    out = tmp_path / f"rt.{fmt}"
    await document_create(DocumentCreateParams(
        path=str(out), format=fmt, title="Report", blocks=_BLOCKS, sheets=_SHEETS,
    ))
    r = await document_read(DocumentReadParams(path=str(out)))
    assert r.success, r.error
    # Numeric/table content survives the round-trip.
    assert "15" in (r.output or "")


@pytest.mark.asyncio
async def test_read_missing_file(tmp_path):
    r = await document_read(DocumentReadParams(path=str(tmp_path / "nope.xlsx")))
    assert not r.success
    assert "not found" in (r.error or "").lower()


@pytest.mark.asyncio
async def test_read_truncates(tmp_path):
    out = tmp_path / "big.txt"
    out.write_text("x" * 5000, encoding="utf-8")
    r = await document_read(DocumentReadParams(path=str(out), max_chars=100))
    assert r.success
    assert r.metadata.get("truncated") is True
    assert len(r.output) < 5000


@pytest.mark.asyncio
async def test_empty_path_rejected(tmp_path):
    r = await document_create(DocumentCreateParams(path="", format="pdf"))
    assert not r.success


@pytest.mark.asyncio
async def test_missing_library_reports_gracefully(tmp_path, monkeypatch):
    # Simulate the optional renderer raising ImportError -> graceful message,
    # never a crash.
    import rune.capabilities.document as docmod

    def _boom(_p, _params):
        raise ImportError("no module")

    monkeypatch.setitem(docmod._RENDERERS, "pdf", (_boom, "fpdf2"))
    r = await document_create(DocumentCreateParams(
        path=str(tmp_path / "x.pdf"), format="pdf", title="t",
    ))
    assert not r.success
    assert "fpdf2" in (r.error or "")


@pytest.mark.asyncio
async def test_registered_in_registry():
    from rune.capabilities.registry import get_capability_registry

    reg = get_capability_registry()
    probe = Path.home() / ".rune_doctest_probe.html"
    try:
        r = await reg.execute("document_create", {
            "path": str(probe), "format": "html", "title": "x",
        })
        assert "Unknown capability" not in (r.error or "")
        assert r.success
    finally:
        probe.unlink(missing_ok=True)
